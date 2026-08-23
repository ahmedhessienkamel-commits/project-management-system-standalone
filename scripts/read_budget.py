from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/ممم.xlsx')
wb = load_workbook(path, data_only=False, read_only=True)
for ws in wb.worksheets:
    print(f'[{ws.title}]')
    for row in ws.iter_rows(values_only=True):
        values = [v for v in row if v is not None]
        if values:
            print('\t'.join(str(v) for v in values))
